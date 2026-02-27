"""
SIENGE EXTRACTOR - VERSÃO PRODUÇÃO
Código limpo e otimizado para uso em produção
"""

import base64
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, List

import pandas as pd
import numpy as np
import requests


DEFAULT_CREDENTIALS_FILE = Path(__file__).with_name('sienge_credentials.json')


def carregar_credenciais(credentials_path: Optional[str] = None) -> Tuple[Optional[str], Optional[str]]:
    """Carrega credenciais do arquivo JSON"""
    path = Path(credentials_path) if credentials_path else DEFAULT_CREDENTIALS_FILE
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
            return data.get('username'), data.get('password')
        except json.JSONDecodeError as exc:
            print(f'Erro ao ler credenciais: {exc}')
            return None, None
    return None, None


class SiengeExtractor:
    """Extrator completo e otimizado para API Sienge"""
    
    def __init__(self, username: str, password: str):
        self.username = username
        self.password = password
        self.base_url = 'https://api.sienge.com.br/trust/public/api'
        
        credentials = f"{username}:{password}"
        encoded = base64.b64encode(credentials.encode('utf-8')).decode('ascii')
        
        self.headers = {
            'Authorization': f'Basic {encoded}',
            'Content-Type': 'application/json',
            'Accept': 'application/json',
        }
        self.cache_dir = Path('cache_sienge')
        self.cache_dir.mkdir(exist_ok=True)
    
    def _fazer_requisicao(self, url: str, params: dict = None, max_tentativas: int = 3) -> dict:
        """Faz requisição HTTP com retry logic"""
        for tentativa in range(max_tentativas):
            try:
                response = requests.get(url, headers=self.headers, params=params, timeout=30)
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 429:
                    time.sleep(60)
                else:
                    if tentativa < max_tentativas - 1:
                        time.sleep(5)
            except requests.exceptions.RequestException:
                if tentativa < max_tentativas - 1:
                    time.sleep(5)
        return {}
    
    def _paginar_resultados(self, url: str, params: dict) -> List[dict]:
        """Paginação robusta"""
        todos_resultados = []
        offset = 0
        offsets_visitados = set()
        tentativas_sem_dados = 0
        
        while tentativas_sem_dados < 3:
            if offset in offsets_visitados:
                break
            offsets_visitados.add(offset)
            
            params['offset'] = offset
            dados = self._fazer_requisicao(url, params)
            
            if not dados:
                tentativas_sem_dados += 1
                offset += 200
                continue
            
            results = dados.get('results', [])
            if not results:
                break
            
            todos_resultados.extend(results)
            tentativas_sem_dados = 0
            
            if len(results) < params.get('limit', 200):
                break
            
            next_offset = dados.get('nextOffset')
            if next_offset is not None and next_offset != offset:
                offset = next_offset
            else:
                offset += len(results)
            
            time.sleep(0.3)
        
        return todos_resultados
    
    def buscar_credores(self, forcar_atualizacao: bool = False) -> pd.DataFrame:
        """Busca credores com cache de 24h"""
        cache_file = self.cache_dir / 'credores.parquet'
        
        if not forcar_atualizacao and cache_file.exists():
            idade_horas = (time.time() - cache_file.stat().st_mtime) / 3600
            if idade_horas < 24:
                return pd.read_parquet(cache_file)
        
        url = f"{self.base_url}/v1/creditors"
        params = {'limit': 200, 'offset': 0}
        todos_credores = self._paginar_resultados(url, params)
        
        if not todos_credores:
            return pd.DataFrame()
        
        df = pd.DataFrame([{
            'supplierId': c.get('id'),
            'nome_fornecedor': c.get('name', ''),
            'nome_fantasia': c.get('tradeName', ''),
            'cpf_cnpj': c.get('cnpj', c.get('cpf', ''))
        } for c in todos_credores])
        
        df.to_parquet(cache_file)
        return df
    
    def listar_empreendimentos(self) -> pd.DataFrame:
        """Lista empreendimentos com filtros aplicados"""
        url = f"{self.base_url}/v1/enterprises"
        params = {'limit': 200, 'offset': 0, 'onlyBuildingsEnabledForIntegration': False}
        
        todos = self._paginar_resultados(url, params)
        if not todos:
            return pd.DataFrame()
        
        df = pd.DataFrame(todos)
        df = df.rename(columns={'id': 'ID', 'name': 'Obra'})
        df['tabela_orcamento'] = df.get('costDatabaseId', 'SEM_TABELA').fillna('SEM_TABELA')
        
        # Filtros - Incluir obras com padrões específicos OU que contenham Nola
        filtro_padroes = df['Obra'].str.contains('SPE|SCP|HLT|HPB|INVESTCORP|NEOON', case=False, na=False)
        filtro_nola = df['Obra'].str.contains('Nola', case=False, na=False)
        df = df[filtro_padroes | filtro_nola]  # OU lógico - pega qualquer um dos dois
        
        df = df[df['Obra'].str.contains('- Obra', case=False, na=False)]
        
        exclusoes = [
            'tabela', 'NÃO USAR', 'NÃO usar', 'não usar', 
            'JN NEGOCIOS', 'JN Negocios',
            'antiga', 'Antiga', 
            'Mare di Capri', 'Costa Club', 
            ' Leme', ' Level', 'Latitude', 'Ocean',
            ' Lume', ' Soho', ' Sollus', ' Unik', 
            'Urban295', 'UNIQUE Estudios', 'INVESTCORP UFV'
        ]
        padrao_exclusao = '|'.join(exclusoes)
        df = df[~df['Obra'].str.contains(padrao_exclusao, case=False, na=False)]
        
        df = df.sort_values('Obra').reset_index(drop=True)
        return df
    
    def extrair_orcamento(self, building_id: int) -> pd.DataFrame:
        """Extrai orçamento de uma obra"""
        url = f"{self.base_url}/bulk-data/v1/building/resources"
        params = {
            'buildingId': building_id,
            'startDate': '2024-01-01',
            'endDate': '2026-12-31',
            'includeDisbursement': False,
            'bdi': '0.00',
            'laborBurden': '0.00'
        }
        
        data = self._fazer_requisicao(url, params)
        if not data or not data.get('data'):
            return pd.DataFrame()
        
        recursos = [{
            'id_insumo': str(item.get('id', '')),
            'codigo_insumo': item.get('resourceCode'),
            'insumo': item.get('description'),
            'categoria': item.get('category', ''),
            'grupo_insumo': item.get('resourceGroup', ''),
            'preco_unitario_orcado': item.get('unitPrice', 0),
            'unidade_medida_orcamento': item.get('unitOfMeasure'),
            'quantidade_orcada': sum(ci.get('quantity', 0) for ci in item.get('buildingCostEstimationItems', [])),
            'valor_total_orcado': sum(ci.get('totalPrice', 0) for ci in item.get('buildingCostEstimationItems', [])),
        } for item in data['data']]
        
        df = pd.DataFrame(recursos)
        for col in ['preco_unitario_orcado', 'quantidade_orcada', 'valor_total_orcado']:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        return df
    
    def extrair_pedidos(self, building_id: int) -> pd.DataFrame:
        """Extrai pedidos de uma obra"""
        url = f"{self.base_url}/v1/purchase-orders"
        params = {
            'buildingId': building_id,
            'startDate': '2024-01-01',
            'endDate': '2026-12-31',
            'consistency': 'CONSISTENT',
            'limit': 200,
            'offset': 0
        }
        
        todos = self._paginar_resultados(url, params)
        if not todos:
            return pd.DataFrame()
        
        pedidos_unicos = {p['id']: p for p in todos}.values()
        df = pd.DataFrame(pedidos_unicos)
        
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        return df
    
    def extrair_itens_pedidos(self, pedidos_ids: list) -> pd.DataFrame:
        """Extrai itens de múltiplos pedidos"""
        todos_itens = []
        
        for pedido_id in pedidos_ids:
            url = f"{self.base_url}/v1/purchase-orders/{pedido_id}/items"
            params = {'limit': 200, 'offset': 0}
            
            itens = self._paginar_resultados(url, params)
            
            for item in itens:
                item['pedido_id'] = pedido_id
                item['detailId'] = item.get('detailId', '')
                item['detailDescription'] = item.get('detailDescription', '')
                item['itemNumber'] = item.get('itemNumber', '')
                item['resourceReference'] = item.get('resourceReference', '')
                item['trademarkDescription'] = item.get('trademarkDescription', '')
            
            todos_itens.extend(itens)
            time.sleep(0.3)
        
        if not todos_itens:
            return pd.DataFrame()
        
        return pd.DataFrame(todos_itens)
    
    def criar_analise_completa(self, df_orcamento, df_pedidos, df_itens, df_credores=None):
        """Cria análise completa com médias por tabela e por obra"""
        if df_itens.empty:
            return pd.DataFrame()
        
        itens_todos = df_itens.copy()
        
        # Normalização
        itens_todos['id_insumo'] = itens_todos.get('resourceId', itens_todos.get('resourceCode', '')).astype(str)
        itens_todos['quantidade_pedido'] = pd.to_numeric(itens_todos.get('quantity', 0), errors='coerce').fillna(0)
        itens_todos['preco_unitario_pedido'] = pd.to_numeric(itens_todos.get('unitPrice', 0), errors='coerce').fillna(0)
        itens_todos['unidade_medida_pedido'] = itens_todos.get('unitOfMeasure', '').str.strip().str.upper()
        itens_todos['detalhe_id'] = itens_todos.get('detailId', '').astype(str).fillna('')
        
        # Calcular médias por tabela
        stats_por_tabela = itens_todos.groupby([
            'tabela_orcamento', 'id_insumo', 'unidade_medida_pedido', 'detalhe_id'
        ])['preco_unitario_pedido'].agg(['mean', 'max', 'min', 'count']).reset_index()
        
        stats_por_tabela.columns = [
            'tabela_orcamento', 'id_insumo', 'unidade_medida_pedido', 'detalhe_id',
            'preco_medio_tabela', 'preco_maximo_tabela', 'preco_minimo_tabela', 'qtd_compras_tabela'
        ]
        
        # Processar cada obra
        obras_unicas = df_itens['id_obra'].unique()
        resultados = []
        
        for obra_id in obras_unicas:
            df_orc = df_orcamento[df_orcamento['id_obra'] == obra_id].copy() if not df_orcamento.empty else pd.DataFrame()
            df_ped = df_pedidos[df_pedidos['id_obra'] == obra_id].copy() if not df_pedidos.empty else pd.DataFrame()
            df_itens_obra = df_itens[df_itens['id_obra'] == obra_id].copy()
            
            resultado = self._processar_obra_individual(
                df_orc, df_itens_obra, df_ped, df_credores, stats_por_tabela
            )
            
            if not resultado.empty:
                resultados.append(resultado)
        
        if not resultados:
            return pd.DataFrame()
        
        return pd.concat(resultados, ignore_index=True)
    
    def _processar_obra_individual(self, df_orcamento_obra, df_itens_obra, df_pedidos_obra, df_credores, stats_por_tabela):
        """Processa uma obra isoladamente"""
        if df_itens_obra.empty:
            return pd.DataFrame()
        
        itens = df_itens_obra.copy()
        itens['id_insumo'] = itens.get('resourceId', itens.get('resourceCode', '')).astype(str)
        
        # Merge com pedidos
        if not df_pedidos_obra.empty:
            pedidos_info = df_pedidos_obra[['id', 'date', 'authorized', 'buyerId', 'supplierId']].copy()
            pedidos_info['pedido_id'] = pedidos_info['id']
            itens = itens.merge(pedidos_info, on='pedido_id', how='left', suffixes=('', '_p'))
        
        # Campos principais
        itens['numero_pedido'] = itens['pedido_id'].astype(str)
        itens['insumo'] = itens.get('resourceDescription', '')
        itens['data_pedido'] = itens.get('date', '')
        itens['pedido_autorizado'] = itens.get('authorized', False)
        itens['buyerId'] = itens.get('buyerId', '')
        itens['supplierId'] = itens.get('supplierId', '')
        itens['quantidade_pedido'] = pd.to_numeric(itens.get('quantity', 0), errors='coerce').fillna(0)
        itens['preco_unitario_pedido'] = pd.to_numeric(itens.get('unitPrice', 0), errors='coerce').fillna(0)
        itens['unidade_medida_pedido'] = itens.get('unitOfMeasure', '').str.strip().str.upper()
        
        # Campos de detalhe
        itens['detalhe_id'] = itens.get('detailId', '').astype(str).fillna('')
        itens['detalhe_descricao'] = itens.get('detailDescription', '').fillna('SEM DETALHE')
        itens['numero_item'] = itens.get('itemNumber', '')
        itens['referencia_insumo'] = itens.get('resourceReference', '')
        itens['marca'] = itens.get('trademarkDescription', '')
        
        # Merge com fornecedores
        if df_credores is not None and not df_credores.empty:
            itens['supplierId'] = pd.to_numeric(itens['supplierId'], errors='coerce')
            df_credores_temp = df_credores.copy()
            df_credores_temp['supplierId'] = pd.to_numeric(df_credores_temp['supplierId'], errors='coerce')
            itens = itens.merge(df_credores_temp[['supplierId', 'nome_fornecedor']], on='supplierId', how='left')
        else:
            itens['nome_fornecedor'] = ''
        
        # Remover duplicatas
        itens = itens.drop_duplicates(
            subset=['id_insumo', 'pedido_id', 'unidade_medida_pedido', 'detalhe_id'], 
            keep='first'
        )
        
        # Adicionar médias por tabela
        itens = itens.merge(
            stats_por_tabela, 
            on=['tabela_orcamento', 'id_insumo', 'unidade_medida_pedido', 'detalhe_id'], 
            how='left'
        )
        
        # Calcular médias por obra
        stats_por_obra = itens.groupby([
            'tabela_orcamento', 'id_insumo', 'unidade_medida_pedido', 'detalhe_id'
        ])['preco_unitario_pedido'].agg(['mean', 'count']).reset_index()
        
        stats_por_obra.columns = [
            'tabela_orcamento', 'id_insumo', 'unidade_medida_pedido', 'detalhe_id',
            'preco_medio_obra', 'qtd_compras_obra'
        ]
        
        itens = itens.merge(
            stats_por_obra, 
            on=['tabela_orcamento', 'id_insumo', 'unidade_medida_pedido', 'detalhe_id'], 
            how='left'
        )
        
        # Merge com orçamento
        if not df_orcamento_obra.empty:
            df_orcamento_obra['unidade_medida_orcamento'] = df_orcamento_obra['unidade_medida_orcamento'].str.strip().str.upper()
            df_orcamento_obra = df_orcamento_obra.drop_duplicates(subset='id_insumo', keep='first')
            
            orcamento_dict = df_orcamento_obra.set_index('id_insumo')[
                ['preco_unitario_orcado', 'quantidade_orcada', 'valor_total_orcado', 
                 'codigo_insumo', 'insumo', 'unidade_medida_orcamento', 'categoria', 'grupo_insumo']
            ].to_dict('index')
            
            itens['preco_unitario_orcado'] = itens['id_insumo'].map(
                lambda x: orcamento_dict.get(x, {}).get('preco_unitario_orcado', np.nan)
            )
            itens['quantidade_orcada'] = itens['id_insumo'].map(
                lambda x: orcamento_dict.get(x, {}).get('quantidade_orcada', np.nan)
            )
            itens['valor_total_orcado'] = itens['id_insumo'].map(
                lambda x: orcamento_dict.get(x, {}).get('valor_total_orcado', np.nan)
            )
            itens['unidade_medida_orcamento'] = itens['id_insumo'].map(
                lambda x: orcamento_dict.get(x, {}).get('unidade_medida_orcamento', '')
            )
            itens['categoria'] = itens['id_insumo'].map(
                lambda x: orcamento_dict.get(x, {}).get('categoria', '')
            )
            itens['grupo_insumo'] = itens['id_insumo'].map(
                lambda x: orcamento_dict.get(x, {}).get('grupo_insumo', '')
            )
            
            itens['insumo'] = itens.apply(
                lambda row: orcamento_dict.get(row['id_insumo'], {}).get('insumo', '') if not row['insumo'] else row['insumo'],
                axis=1
            )
            
            # Status orçamento
            itens['status_orcamento'] = 'Sem Orcamento'
            itens.loc[itens['preco_unitario_orcado'] > 0, 'status_orcamento'] = 'Com Orcamento'
            itens.loc[itens['preco_unitario_orcado'].isna(), 'status_orcamento'] = 'Sem Valor'
            
            # Variações
            mask_unidade_igual = (
                (itens['unidade_medida_orcamento'] == itens['unidade_medida_pedido']) &
                (itens['preco_unitario_orcado'] > 0) & 
                (itens['preco_unitario_pedido'].notna())
            )
            
            itens['variacao_orcado_pedido_valor'] = np.where(
                mask_unidade_igual,
                itens['preco_unitario_pedido'] - itens['preco_unitario_orcado'],
                np.nan
            )
            
            itens['variacao_orcado_pedido_pct'] = np.where(
                mask_unidade_igual,
                itens['variacao_orcado_pedido_valor'] / itens['preco_unitario_orcado'],
                np.nan
            )
            
            itens['variacao_pedido_media_tabela_valor'] = (
                itens['preco_unitario_pedido'] - itens['preco_medio_tabela']
            )
            itens['variacao_pedido_media_tabela_pct'] = np.where(
                itens['preco_medio_tabela'] > 0,
                itens['variacao_pedido_media_tabela_valor'] / itens['preco_medio_tabela'],
                np.nan
            )
            
            itens['variacao_pedido_media_obra_valor'] = (
                itens['preco_unitario_pedido'] - itens['preco_medio_obra']
            )
            itens['variacao_pedido_media_obra_pct'] = np.where(
                itens['preco_medio_obra'] > 0,
                itens['variacao_pedido_media_obra_valor'] / itens['preco_medio_obra'],
                np.nan
            )
        else:
            itens['preco_unitario_orcado'] = np.nan
            itens['quantidade_orcada'] = np.nan
            itens['valor_total_orcado'] = np.nan
            itens['unidade_medida_orcamento'] = ''
            itens['categoria'] = ''
            itens['grupo_insumo'] = ''
            itens['status_orcamento'] = 'Sem Orcamento'
            
            for col in ['variacao_orcado_pedido_valor', 'variacao_orcado_pedido_pct',
                       'variacao_pedido_media_tabela_valor', 'variacao_pedido_media_tabela_pct',
                       'variacao_pedido_media_obra_valor', 'variacao_pedido_media_obra_pct']:
                itens[col] = np.nan
        
        # Colunas finais
        colunas_finais = [
            'tabela_orcamento', 'id_obra', 'nome_obra', 
            'id_insumo', 'insumo', 'categoria', 'grupo_insumo',
            'numero_pedido', 'numero_item', 'data_pedido', 'pedido_autorizado',
            'detalhe_id', 'detalhe_descricao', 'referencia_insumo', 'marca',
            'supplierId', 'nome_fornecedor', 'buyerId',
            'quantidade_orcada', 'preco_unitario_orcado', 'unidade_medida_orcamento', 'status_orcamento',
            'quantidade_pedido', 'preco_unitario_pedido', 'unidade_medida_pedido',
            'preco_medio_tabela', 'preco_medio_obra', 
            'qtd_compras_tabela', 'qtd_compras_obra',
            'preco_maximo_tabela', 'preco_minimo_tabela',
            'variacao_orcado_pedido_valor', 'variacao_orcado_pedido_pct',
            'variacao_pedido_media_tabela_valor', 'variacao_pedido_media_tabela_pct',
            'variacao_pedido_media_obra_valor', 'variacao_pedido_media_obra_pct',
        ]
        
        for col in colunas_finais:
            if col not in itens.columns:
                itens[col] = '' if col in ['tabela_orcamento', 'id_obra', 'nome_obra', 'insumo', 
                                           'categoria', 'grupo_insumo', 'detalhe_descricao', 
                                           'marca', 'referencia_insumo'] else np.nan
        
        resultado = itens[colunas_finais]
        
        # Remover duplicatas finais
        resultado = resultado.drop_duplicates(
            subset=['id_insumo', 'numero_pedido', 'detalhe_id', 'unidade_medida_pedido'],
            keep='first'
        )
        
        return resultado
    
    def _aplicar_formatacao_tabela_excel(self, arquivo: str, nome_tabela: str = "TabelaAnalise"):
        """Aplica formatação de tabela Excel a um arquivo"""
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(arquivo)
        ws = wb.active
        
        # Definir o intervalo da tabela
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row <= 1:  # Apenas cabeçalho ou vazio
            wb.close()
            return
        
        ultima_coluna = get_column_letter(max_col)
        
        # Criar tabela do Excel
        ref = f"A1:{ultima_coluna}{max_row}"
        tab = Table(displayName=nome_tabela, ref=ref)
        
        # Estilo da tabela
        style = TableStyleInfo(
            name="TableStyleMedium9",  
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        
        # Adicionar tabela à planilha
        ws.add_table(tab)
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primeira linha (cabeçalho)
        ws.freeze_panes = 'A2'
        
        # Salvar com formatação
        wb.save(arquivo)
        wb.close()
    
    def salvar_resultado(self, df_final: pd.DataFrame, timestamp: str) -> str:
        """Salva resultado em múltiplos formatos com formatação de tabela Excel"""
        
        # ============================================
        # ARQUIVO 1: COMPLETO
        # ============================================
        base_name = 'Analise_Orcamento_Pedidos'
        arquivo_completo = f'{base_name}.xlsx'
        
        print(f"\n  Salvando arquivo completo: {arquivo_completo}")
        df_final.to_excel(arquivo_completo, index=False, engine='openpyxl')
        self._aplicar_formatacao_tabela_excel(arquivo_completo, "TabelaCompleta")
        
        # ============================================
        # ARQUIVO 2: APENAS PEDIDOS DESAUTORIZADOS (VALORES ÚNICOS)
        # ============================================
        arquivo_desautorizado = 'Analise_Pedidos_Desautorizados.xlsx'
        
        # Filtrar apenas pedidos não autorizados
        df_desautorizados = df_final[df_final['pedido_autorizado'] == False].copy()
        
        if not df_desautorizados.empty:
            # Remover duplicatas por numero_pedido (mantém primeira ocorrência)
            df_desautorizados = df_desautorizados.drop_duplicates(
                subset=['numero_pedido'], 
                keep='first'
            )
            
            print(f"  Salvando arquivo desautorizados: {arquivo_desautorizado}")
            print(f"    - {len(df_desautorizados)} pedidos desautorizados únicos")
            
            df_desautorizados.to_excel(arquivo_desautorizado, index=False, engine='openpyxl')
            self._aplicar_formatacao_tabela_excel(arquivo_desautorizado, "TabelaDesautorizados")
        else:
            print(f"  ⚠ Nenhum pedido desautorizado encontrado")
        
        # ============================================
        # OUTROS FORMATOS (apenas completo)
        # ============================================
        df_final.to_parquet(f'{base_name}.parquet', compression='snappy')
        df_final.to_csv(f'{base_name}.csv', index=False, encoding='utf-8-sig')
        
        return base_name


if __name__ == "__main__":
    print("="*70)
    print("EXTRAÇÃO SIENGE - PRODUÇÃO")
    print("="*70)
    
    username, password = carregar_credenciais()
    if not username or not password:
        print("\n⚠ ERRO: Configure sienge_credentials.json")
        exit(1)
    
    inicio = time.time()
    extrator = SiengeExtractor(username, password)
    
    print("\n[1/5] Buscando credores...")
    df_credores = extrator.buscar_credores()
    print(f"✓ {len(df_credores)} credores")
    
    print("\n[2/5] Listando empreendimentos...")
    df_empreendimentos = extrator.listar_empreendimentos()
    print(f"✓ {len(df_empreendimentos)} obras selecionadas")
    
    if df_empreendimentos.empty:
        print("\n⚠ Nenhum empreendimento encontrado")
        exit(1)
    
    print("\n[3/5] Extraindo dados das obras...")
    todos_orc, todos_ped, todos_itens = [], [], []
    
    for idx, row in df_empreendimentos.iterrows():
        obra_id, obra_nome, tabela = row['ID'], row['Obra'], row['tabela_orcamento']
        print(f"  Processando {idx+1}/{len(df_empreendimentos)}: {obra_nome[:40]}...")
        
        try:
            df_orc = extrator.extrair_orcamento(obra_id)
            if not df_orc.empty:
                df_orc['id_obra'] = obra_id
                df_orc['nome_obra'] = obra_nome
                df_orc['tabela_orcamento'] = tabela
                todos_orc.append(df_orc)
            
            df_ped = extrator.extrair_pedidos(obra_id)
            if not df_ped.empty:
                df_ped['id_obra'] = obra_id
                df_ped['nome_obra'] = obra_nome
                df_ped['tabela_orcamento'] = tabela
                todos_ped.append(df_ped)
                
                ids = df_ped['id'].tolist()
                df_it = extrator.extrair_itens_pedidos(ids)
                if not df_it.empty:
                    df_it['id_obra'] = obra_id
                    df_it['nome_obra'] = obra_nome
                    df_it['tabela_orcamento'] = tabela
                    todos_itens.append(df_it)
        except Exception as e:
            print(f"    ⚠ Erro: {str(e)[:50]}")
            continue
    
    if not todos_itens:
        print("\n⚠ Nenhum item coletado")
        exit(1)
    
    print("\n[4/5] Consolidando e calculando médias...")
    df_orcamento_consolidado = pd.concat(todos_orc, ignore_index=True) if todos_orc else pd.DataFrame()
    df_pedidos_consolidado = pd.concat(todos_ped, ignore_index=True) if todos_ped else pd.DataFrame()
    df_itens_consolidado = pd.concat(todos_itens, ignore_index=True)
    
    df_final = extrator.criar_analise_completa(
        df_orcamento_consolidado,
        df_pedidos_consolidado,
        df_itens_consolidado,
        df_credores
    )
    
    if df_final is None or df_final.empty:
        print("\n⚠ Nenhum resultado gerado")
        exit(1)
    
    print("\n[5/5] Salvando resultados...")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = extrator.salvar_resultado(df_final, timestamp)
    
    tempo_total = time.time() - inicio
    
    print(f"\n{'='*70}")
    print(f"EXTRAÇÃO CONCLUÍDA")
    print(f"{'='*70}")
    print(f"Tempo: {tempo_total/60:.1f} min")
    print(f"Obras: {df_final['id_obra'].nunique()}")
    print(f"Pedidos: {df_final['numero_pedido'].nunique()}")
    print(f"Linhas totais: {len(df_final):,}")
    print(f"\n📁 ARQUIVOS GERADOS:")
    print(f"  1. {base_name}.xlsx (COMPLETO - todas as linhas)")
    print(f"  2. Analise_Pedidos_Desautorizados.xlsx (FILTRADO)")
    print(f"  3. {base_name}.parquet")
    print(f"  4. {base_name}.csv")
    print(f"{'='*70}")